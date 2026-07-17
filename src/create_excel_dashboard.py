"""YES24 IT 베스트셀러 CSV → Excel 대시보드 생성 스크립트."""

import os
import re
from collections import Counter

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR = os.path.join(BASE_DIR, "data")
INPUT_CSV = os.path.join(DATA_DIR, "yes24_it_bestsellers_with_desc.csv")
OUTPUT_XLSX = os.path.join(DATA_DIR, "yes24_it_bestsellers_dashboard.xlsx")

df = pd.read_csv(INPUT_CSV)
df["Rank"] = pd.to_numeric(df["Rank"], errors="coerce").fillna(999).astype(int)
df["Price"] = pd.to_numeric(df["Price"], errors="coerce").fillna(0).astype(int)
df["Rating"] = pd.to_numeric(df["Rating"], errors="coerce").fillna(0.0)
df["ReviewCount"] = pd.to_numeric(df["ReviewCount"], errors="coerce").fillna(0).astype(int)

def parse_ym(s):
    if not isinstance(s, str):
        return "", ""
    m = re.search(r"(\d{4})년\s*(\d{2})월", s)
    if m:
        return m.group(1), m.group(2)
    m2 = re.search(r"(\d{4})년", s)
    if m2:
        return m2.group(1), "01"
    return "", ""

ym = df["PublishDate"].apply(parse_ym)
df["출판연도"] = [y for y, _ in ym]
df["출판월"] = [m for _, m in ym]
df["연월"] = df.apply(lambda r: f"{r['출판연도']}-{r['출판월']}" if r["출판연도"] else "", axis=1)

DARK_BLUE = "1A1A2E"
ACCENT_BLUE = "0F3460"
LIGHT_BG = "F0F4F8"
WHITE = "FFFFFF"
HEADER_FILL = PatternFill("solid", fgColor=DARK_BLUE)
HEADER_FONT = Font(name="Arial", bold=True, color=WHITE, size=11)
KPI_FILL = PatternFill("solid", fgColor=LIGHT_BG)
KPI_FONT = Font(name="Arial", bold=True, color=DARK_BLUE, size=14)
KPI_LABEL_FONT = Font(name="Arial", color="555555", size=10)
TITLE_FONT = Font(name="Arial", bold=True, color=DARK_BLUE, size=16)
SUBTITLE_FONT = Font(name="Arial", bold=True, color=ACCENT_BLUE, size=12)
DATA_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
CENTER = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_data(ws, r1, r2, ncol):
    alt = PatternFill("solid", fgColor="F8FAFC")
    for r in range(r1, r2 + 1):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")
            if (r - r1) % 2 == 1:
                cell.fill = alt


wb = Workbook()

# ===================== Sheet 1: 도서 목록 ====================================
ws_list = wb.active
ws_list.title = "도서 목록"
ws_list.sheet_properties.tabColor = "0F3460"

headers = ["순위", "제목", "저자", "출판사", "출판일", "가격", "평점", "리뷰 수", "링크", "소개"]
for c, h in enumerate(headers, 1):
    ws_list.cell(row=1, column=c, value=h)
style_header(ws_list, 1, len(headers))

for i, (_, row) in enumerate(df.iterrows(), 2):
    ws_list.cell(row=i, column=1, value=row["Rank"])
    ws_list.cell(row=i, column=2, value=row["Title"])
    ws_list.cell(row=i, column=3, value=row["Author"])
    ws_list.cell(row=i, column=4, value=row["Publisher"])
    ws_list.cell(row=i, column=5, value=row["PublishDate"])
    ws_list.cell(row=i, column=6, value=row["Price"])
    ws_list.cell(row=i, column=7, value=row["Rating"])
    ws_list.cell(row=i, column=8, value=row["ReviewCount"])
    ws_list.cell(row=i, column=9, value=row["Link"])
    ws_list.cell(row=i, column=10, value=row["Description"] if pd.notna(row["Description"]) else "")

LAST = len(df) + 1
style_data(ws_list, 2, LAST, len(headers))
for r in range(2, LAST + 1):
    ws_list.cell(row=r, column=1).alignment = CENTER
    ws_list.cell(row=r, column=6).number_format = "#,##0"
    ws_list.cell(row=r, column=7).number_format = "0.0"
    ws_list.cell(row=r, column=8).number_format = "#,##0"

widths = [6, 45, 25, 18, 14, 10, 8, 10, 40, 60]
for c, w in enumerate(widths, 1):
    ws_list.column_dimensions[get_column_letter(c)].width = w
ws_list.freeze_panes = "A2"
ws_list.auto_filter.ref = f"A1:J{LAST}"

# ===================== Sheet 2: 대시보드 ====================================
ws = wb.create_sheet("대시보드", 0)
ws.sheet_properties.tabColor = DARK_BLUE

# ── 타이틀 ──
ws.merge_cells("A1:J1")
ws["A1"] = "YES24 IT 베스트셀러 대시보드"
ws["A1"].font = TITLE_FONT
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 40

ws.merge_cells("A2:J2")
ws["A2"] = "IT/모바일 분야 베스트셀러 1,000권 탐색적 데이터 분석"
ws["A2"].font = SUBTITLE_FONT
ws["A2"].alignment = Alignment(horizontal="center")
ws.row_dimensions[2].height = 25

# ── KPI (Row 4-5) ──
kpi_data = [
    ("A", "총 도서 수", f"=COUNTA('도서 목록'!A2:A{LAST})", "#,##0"),
    ("C", "평균 평점", f"=AVERAGE('도서 목록'!F2:F{LAST})", "0.00"),
    ("E", "평균 가격", f"=AVERAGE('도서 목록'!E2:E{LAST})", "#,##0"),
    ("G", "총 리뷰 수", f"=SUM('도서 목록'!G2:G{LAST})", "#,##0"),
    ("I", "출판사 수", f"=COUNTA('도서 목록'!C2:C{LAST})-COUNTBLANK('도서 목록'!C2:C{LAST})", "#,##0"),
]
for col, label, formula, fmt in kpi_data:
    c1 = ws[f"{col}4"]
    c1.value = label
    c1.font = KPI_LABEL_FONT
    c1.alignment = CENTER
    c1.fill = KPI_FILL
    c1.border = THIN_BORDER
    c2 = ws[f"{col}5"]
    c2.value = formula
    c2.font = KPI_FONT
    c2.alignment = CENTER
    c2.fill = KPI_FILL
    c2.border = THIN_BORDER
    c2.number_format = fmt
    # merged-style adjacent cells
    nxt = chr(ord(col) + 1)
    for r in ["4", "5"]:
        ws[f"{nxt}{r}"].fill = KPI_FILL
        ws[f"{nxt}{r}"].border = THIN_BORDER

ws.row_dimensions[4].height = 22
ws.row_dimensions[5].height = 35

# ── 차트 데이터 영역 (Row 7부터, A-B: 출판사 Top10, D-E: 가격 분포, G-H: 연도 동향, J-K: 평점 분포) ──

# 출판사 Top 10
pub_counts = df["Publisher"].value_counts().head(10)
ws.cell(row=7, column=1, value="출판사").font = HEADER_FONT
ws.cell(row=7, column=1).fill = HEADER_FILL
ws.cell(row=7, column=1).border = THIN_BORDER
ws.cell(row=7, column=2, value="도서 수").font = HEADER_FONT
ws.cell(row=7, column=2).fill = HEADER_FILL
ws.cell(row=7, column=2).border = THIN_BORDER
for i, (name, cnt) in enumerate(pub_counts.items(), 8):
    ws.cell(row=i, column=1, value=name).font = DATA_FONT
    ws.cell(row=i, column=1).border = THIN_BORDER
    ws.cell(row=i, column=2, value=int(cnt)).font = DATA_FONT
    ws.cell(row=i, column=2).border = THIN_BORDER
    ws.cell(row=i, column=2).number_format = "#,##0"

# 가격 분포
price_bins = [0, 15000, 20000, 25000, 30000, 35000, 70000]
price_labels = ["~1.5만", "1.5~2만", "2~2.5만", "2.5~3만", "3~3.5만", "3.5만~"]
df["가격구간"] = pd.cut(df["Price"], bins=price_bins, labels=price_labels, right=True)
price_dist = df["가격구간"].value_counts().reindex(price_labels).fillna(0).astype(int)

ws.cell(row=7, column=4, value="가격대").font = HEADER_FONT
ws.cell(row=7, column=4).fill = HEADER_FILL
ws.cell(row=7, column=4).border = THIN_BORDER
ws.cell(row=7, column=5, value="도서 수").font = HEADER_FONT
ws.cell(row=7, column=5).fill = HEADER_FILL
ws.cell(row=7, column=5).border = THIN_BORDER
for i, (label, cnt) in enumerate(price_dist.items(), 8):
    ws.cell(row=i, column=4, value=label).font = DATA_FONT
    ws.cell(row=i, column=4).border = THIN_BORDER
    ws.cell(row=i, column=5, value=int(cnt)).font = DATA_FONT
    ws.cell(row=i, column=5).border = THIN_BORDER
    ws.cell(row=i, column=5).number_format = "#,##0"

# 연도별 동향
year_counts = df[df["출판연도"] != ""].groupby("출판연도").size().reset_index(name="도서수").sort_values("출판연도")
ws.cell(row=7, column=7, value="연도").font = HEADER_FONT
ws.cell(row=7, column=7).fill = HEADER_FILL
ws.cell(row=7, column=7).border = THIN_BORDER
ws.cell(row=7, column=8, value="도서 수").font = HEADER_FONT
ws.cell(row=7, column=8).fill = HEADER_FILL
ws.cell(row=7, column=8).border = THIN_BORDER
for i, (_, row) in enumerate(year_counts.iterrows(), 8):
    ws.cell(row=i, column=7, value=row["출판연도"]).font = DATA_FONT
    ws.cell(row=i, column=7).border = THIN_BORDER
    ws.cell(row=i, column=8, value=int(row["도서수"])).font = DATA_FONT
    ws.cell(row=i, column=8).border = THIN_BORDER
    ws.cell(row=i, column=8).number_format = "#,##0"

# 평점 분포
bins_rt = [0, 5, 7, 8, 9, 9.5, 10.1]
labels_rt = ["~5.0", "5.1~7.0", "7.1~8.0", "8.1~9.0", "9.1~9.5", "9.6~10.0"]
df["평점구간"] = pd.cut(df["Rating"], bins=bins_rt, labels=labels_rt, right=True)
rating_dist = df["평점구간"].value_counts().reindex(labels_rt).fillna(0).astype(int)

ws.cell(row=7, column=10, value="평점 구간").font = HEADER_FONT
ws.cell(row=7, column=10).fill = HEADER_FILL
ws.cell(row=7, column=10).border = THIN_BORDER
ws.cell(row=7, column=11, value="도서 수").font = HEADER_FONT
ws.cell(row=7, column=11).fill = HEADER_FILL
ws.cell(row=7, column=11).border = THIN_BORDER
for i, (label, cnt) in enumerate(rating_dist.items(), 8):
    ws.cell(row=i, column=10, value=label).font = DATA_FONT
    ws.cell(row=i, column=10).border = THIN_BORDER
    ws.cell(row=i, column=11, value=int(cnt)).font = DATA_FONT
    ws.cell(row=i, column=11).border = THIN_BORDER
    ws.cell(row=i, column=11).number_format = "#,##0"

# 리뷰 Top 10
top10_rev = df.nlargest(10, "ReviewCount")
ws.cell(row=20, column=1, value="리뷰 Top 10 제목").font = HEADER_FONT
ws.cell(row=20, column=1).fill = HEADER_FILL
ws.cell(row=20, column=1).border = THIN_BORDER
ws.cell(row=20, column=2, value="리뷰 수").font = HEADER_FONT
ws.cell(row=20, column=2).fill = HEADER_FILL
ws.cell(row=20, column=2).border = THIN_BORDER
for i, (_, row) in enumerate(top10_rev.iterrows(), 21):
    t = str(row["Title"])[:30] + ("…" if len(str(row["Title"])) > 30 else "")
    ws.cell(row=i, column=1, value=t).font = DATA_FONT
    ws.cell(row=i, column=1).border = THIN_BORDER
    ws.cell(row=i, column=2, value=int(row["ReviewCount"])).font = DATA_FONT
    ws.cell(row=i, column=2).border = THIN_BORDER
    ws.cell(row=i, column=2).number_format = "#,##0"

# 키워드 Top 15
titles_text = " ".join(df["Title"].dropna().tolist())
words = re.findall(r"[가-힣a-zA-Z]{2,}", titles_text)
stop_words = {"with", "and", "the", "for", "using", "codes", "chatgpt"}
words = [w for w in words if w.lower() not in stop_words]
kw_top15 = Counter(words).most_common(15)

ws.cell(row=20, column=4, value="키워드").font = HEADER_FONT
ws.cell(row=20, column=4).fill = HEADER_FILL
ws.cell(row=20, column=4).border = THIN_BORDER
ws.cell(row=20, column=5, value="빈도").font = HEADER_FONT
ws.cell(row=20, column=5).fill = HEADER_FILL
ws.cell(row=20, column=5).border = THIN_BORDER
for i, (word, cnt) in enumerate(kw_top15, 21):
    ws.cell(row=i, column=4, value=word).font = DATA_FONT
    ws.cell(row=i, column=4).border = THIN_BORDER
    ws.cell(row=i, column=5, value=cnt).font = DATA_FONT
    ws.cell(row=i, column=5).border = THIN_BORDER

# ── 차트 배치 ──
# 1. 출판사 Top 10
c1 = BarChart()
c1.type = "bar"
c1.style = 10
c1.title = "출판사별 도서 수 Top 10"
c1.y_axis.title = "출판사"
c1.x_axis.title = "도서 수"
c1.width = 22
c1.height = 13
d1 = Reference(ws, min_col=2, min_row=7, max_row=17)
ca1 = Reference(ws, min_col=1, min_row=8, max_row=17)
c1.add_data(d1, titles_from_data=True)
c1.set_categories(ca1)
c1.legend = None
ws.add_chart(c1, "A32")

# 2. 가격 분포
c2 = BarChart()
c2.type = "col"
c2.style = 10
c2.title = "가격대별 도서 분포"
c2.y_axis.title = "도서 수"
c2.x_axis.title = "가격대"
c2.width = 22
c2.height = 13
d2 = Reference(ws, min_col=5, min_row=7, max_row=13)
ca2 = Reference(ws, min_col=4, min_row=8, max_row=13)
c2.add_data(d2, titles_from_data=True)
c2.set_categories(ca2)
c2.legend = None
ws.add_chart(c2, "G32")

# 3. 연도별 동향
c3 = LineChart()
c3.style = 10
c3.title = "연도별 출판 동향"
c3.y_axis.title = "도서 수"
c3.x_axis.title = "연도"
c3.width = 22
c3.height = 13
yr_end = 7 + len(year_counts)
d3 = Reference(ws, min_col=8, min_row=7, max_row=yr_end)
ca3 = Reference(ws, min_col=7, min_row=8, max_row=yr_end)
c3.add_data(d3, titles_from_data=True)
c3.set_categories(ca3)
c3.legend = None
ws.add_chart(c3, "A48")

# 4. 평점 분포
c4 = BarChart()
c4.type = "col"
c4.style = 10
c4.title = "평점 구간별 도서 분포"
c4.y_axis.title = "도서 수"
c4.x_axis.title = "평점 구간"
c4.width = 22
c4.height = 13
d4 = Reference(ws, min_col=11, min_row=7, max_row=13)
ca4 = Reference(ws, min_col=10, min_row=8, max_row=13)
c4.add_data(d4, titles_from_data=True)
c4.set_categories(ca4)
c4.legend = None
ws.add_chart(c4, "G48")

# 5. 리뷰 Top 10
c5 = BarChart()
c5.type = "bar"
c5.style = 10
c5.title = "리뷰 수 Top 10 도서"
c5.y_axis.title = "도서"
c5.x_axis.title = "리뷰 수"
c5.width = 22
c5.height = 13
d5 = Reference(ws, min_col=2, min_row=20, max_row=30)
ca5 = Reference(ws, min_col=1, min_row=21, max_row=30)
c5.add_data(d5, titles_from_data=True)
c5.set_categories(ca5)
c5.legend = None
ws.add_chart(c5, "A64")

# 6. 키워드 Top 15
c6 = BarChart()
c6.type = "bar"
c6.style = 10
c6.title = "제목 핵심 키워드 Top 15"
c6.y_axis.title = "키워드"
c6.x_axis.title = "빈도"
c6.width = 22
c6.height = 13
d6 = Reference(ws, min_col=5, min_row=20, max_row=35)
ca6 = Reference(ws, min_col=4, min_row=21, max_row=35)
c6.add_data(d6, titles_from_data=True)
c6.set_categories(ca6)
c6.legend = None
ws.add_chart(c6, "G64")

# 대시보드 열 너비
for c in range(1, 12):
    ws.column_dimensions[get_column_letter(c)].width = 16

# ===================== Sheet 3: 출판사 분석 =================================
ws_pub = wb.create_sheet("출판사 분석")
ws_pub.sheet_properties.tabColor = "16213E"

pub_stats = df.groupby("Publisher").agg(
    도서수=("Title", "count"),
    평균평점=("Rating", "mean"),
    평균가격=("Price", "mean"),
    총리뷰=("ReviewCount", "sum"),
).sort_values("도서수", ascending=False).reset_index()

pub_h = ["출판사", "도서 수", "평균 평점", "평균 가격", "총 리뷰 수"]
for c, h in enumerate(pub_h, 1):
    ws_pub.cell(row=1, column=c, value=h)
style_header(ws_pub, 1, len(pub_h))

for i, (_, row) in enumerate(pub_stats.iterrows(), 2):
    ws_pub.cell(row=i, column=1, value=row["Publisher"])
    ws_pub.cell(row=i, column=2, value=int(row["도서수"]))
    ws_pub.cell(row=i, column=3, value=round(row["평균평점"], 2))
    ws_pub.cell(row=i, column=4, value=int(round(row["평균가격"])))
    ws_pub.cell(row=i, column=5, value=int(row["총리뷰"]))

pub_last = len(pub_stats) + 1
style_data(ws_pub, 2, pub_last, len(pub_h))
for r in range(2, pub_last + 1):
    ws_pub.cell(row=r, column=2).number_format = "#,##0"
    ws_pub.cell(row=r, column=3).number_format = "0.00"
    ws_pub.cell(row=r, column=4).number_format = "#,##0"
    ws_pub.cell(row=r, column=5).number_format = "#,##0"

for c, w in enumerate([22, 10, 12, 12, 12], 1):
    ws_pub.column_dimensions[get_column_letter(c)].width = w
ws_pub.freeze_panes = "A2"
ws_pub.auto_filter.ref = f"A1:E{pub_last}"

# ===================== Sheet 4: 키워드 분석 =================================
ws_kw = wb.create_sheet("키워드 분석")
ws_kw.sheet_properties.tabColor = "7C3AED"

kw_all = Counter(words).most_common(30)
kw_h = ["키워드", "빈도"]
for c, h in enumerate(kw_h, 1):
    ws_kw.cell(row=1, column=c, value=h)
style_header(ws_kw, 1, len(kw_h))

for i, (word, cnt) in enumerate(kw_all, 2):
    ws_kw.cell(row=i, column=1, value=word)
    ws_kw.cell(row=i, column=2, value=cnt)

kw_last = len(kw_all) + 1
style_data(ws_kw, 2, kw_last, len(kw_h))

chart_kw = BarChart()
chart_kw.type = "bar"
chart_kw.title = "제목 핵심 키워드 Top 30"
chart_kw.y_axis.title = "키워드"
chart_kw.x_axis.title = "빈도"
chart_kw.width = 22
chart_kw.height = 16
d_kw = Reference(ws_kw, min_col=2, min_row=1, max_row=kw_last)
ca_kw = Reference(ws_kw, min_col=1, min_row=2, max_row=kw_last)
chart_kw.add_data(d_kw, titles_from_data=True)
chart_kw.set_categories(ca_kw)
chart_kw.legend = None
ws_kw.add_chart(chart_kw, "D1")

for c, w in enumerate([18, 10], 1):
    ws_kw.column_dimensions[get_column_letter(c)].width = w

# ===================== Sheet 5: 평점리뷰 분석 ================================
ws_rt = wb.create_sheet("평점리뷰 분석")
ws_rt.sheet_properties.tabColor = "F59E0B"

rt_h = ["평점 구간", "도서 수"]
for c, h in enumerate(rt_h, 1):
    ws_rt.cell(row=1, column=c, value=h)
style_header(ws_rt, 1, len(rt_h))

for i, (label, cnt) in enumerate(rating_dist.items(), 2):
    ws_rt.cell(row=i, column=1, value=label)
    ws_rt.cell(row=i, column=2, value=int(cnt))

rt_last = len(rating_dist) + 1
style_data(ws_rt, 2, rt_last, len(rt_h))

chart_rt = BarChart()
chart_rt.type = "col"
chart_rt.title = "평점 구간별 도서 분포"
chart_rt.y_axis.title = "도서 수"
chart_rt.x_axis.title = "평점 구간"
chart_rt.width = 18
chart_rt.height = 12
d_rt = Reference(ws_rt, min_col=2, min_row=1, max_row=rt_last)
ca_rt = Reference(ws_rt, min_col=1, min_row=2, max_row=rt_last)
chart_rt.add_data(d_rt, titles_from_data=True)
chart_rt.set_categories(ca_rt)
chart_rt.legend = None
ws_rt.add_chart(chart_rt, "D1")

# 리뷰 Top 20
top20 = df.nlargest(20, "ReviewCount")[["Title", "Author", "Publisher", "Rating", "ReviewCount", "Price"]]
tr_h = ["제목", "저자", "출판사", "평점", "리뷰 수", "가격"]
tr_r = rt_last + 2
for c, h in enumerate(tr_h, 1):
    ws_rt.cell(row=tr_r, column=c, value=h)
style_header(ws_rt, tr_r, len(tr_h))

for i, (_, row) in enumerate(top20.iterrows(), tr_r + 1):
    ws_rt.cell(row=i, column=1, value=row["Title"])
    ws_rt.cell(row=i, column=2, value=row["Author"])
    ws_rt.cell(row=i, column=3, value=row["Publisher"])
    ws_rt.cell(row=i, column=4, value=row["Rating"])
    ws_rt.cell(row=i, column=5, value=int(row["ReviewCount"]))
    ws_rt.cell(row=i, column=6, value=int(row["Price"]))

tr_end = tr_r + len(top20)
style_data(ws_rt, tr_r + 1, tr_end, len(tr_h))
for r in range(tr_r + 1, tr_end + 1):
    ws_rt.cell(row=r, column=4).number_format = "0.0"
    ws_rt.cell(row=r, column=5).number_format = "#,##0"
    ws_rt.cell(row=r, column=6).number_format = "#,##0"

for c, w in enumerate([45, 22, 18, 8, 10, 10], 1):
    ws_rt.column_dimensions[get_column_letter(c)].width = w

# ===================== Sheet 6: 연도월별 동향 ================================
ws_tr = wb.create_sheet("연도월별 동향")
ws_tr.sheet_properties.tabColor = "E11D48"

tr_h2 = ["연도", "도서 수"]
for c, h in enumerate(tr_h2, 1):
    ws_tr.cell(row=1, column=c, value=h)
style_header(ws_tr, 1, len(tr_h2))

for i, (_, row) in enumerate(year_counts.iterrows(), 2):
    ws_tr.cell(row=i, column=1, value=row["출판연도"])
    ws_tr.cell(row=i, column=2, value=int(row["도서수"]))

yr_last = len(year_counts) + 1
style_data(ws_tr, 2, yr_last, len(tr_h2))

# 월별
monthly = df[df["연월"] != ""].groupby("연월").size().reset_index(name="도서수").sort_values("연월")
m_h = ["연월", "도서 수"]
for c, h in enumerate(m_h, 4):
    ws_tr.cell(row=1, column=c, value=h)
style_header(ws_tr, 4, 5)

for i, (_, row) in enumerate(monthly.iterrows(), 2):
    ws_tr.cell(row=i, column=4, value=row["연월"])
    ws_tr.cell(row=i, column=5, value=int(row["도서수"]))

mo_last = max(yr_last, len(monthly) + 1)
style_data(ws_tr, 2, mo_last, 5)

chart_yr = LineChart()
chart_yr.title = "연도별 출판 동향"
chart_yr.y_axis.title = "도서 수"
chart_yr.x_axis.title = "연도"
chart_yr.width = 20
chart_yr.height = 12
d_yr = Reference(ws_tr, min_col=2, min_row=1, max_row=yr_last)
ca_yr = Reference(ws_tr, min_col=1, min_row=2, max_row=yr_last)
chart_yr.add_data(d_yr, titles_from_data=True)
chart_yr.set_categories(ca_yr)
chart_yr.legend = None
ws_tr.add_chart(chart_yr, f"A{yr_last + 2}")

for c, w in enumerate([10, 10, 2, 12, 10], 1):
    ws_tr.column_dimensions[get_column_letter(c)].width = w

# ── 저장 ──
wb.save(OUTPUT_XLSX)
print(f"Dashboard saved: {OUTPUT_XLSX}")
